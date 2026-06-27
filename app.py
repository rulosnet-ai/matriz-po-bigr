"""
Matriz de Producción - Generador automático desde PDF de Orden de Compra
Cliente objetivo: Big R (formato estándar de PO)

Cómo correr:
    streamlit run app.py
"""

import re
import io
from collections import defaultdict

import streamlit as st
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ----------------------------------------------------------------------------
# EXTRACCIÓN DE DATOS DEL PDF
# ----------------------------------------------------------------------------

# Patrón de una línea de detalle, por ejemplo:
# 1 1 1 2129550 M 5 PCKT JEAN MSH 29X30 BR39-AFJNG-MSH 18.60 EA 18.60
LINE_PATTERN = re.compile(
    r"^(?P<line_no>\d+)\s+"
    r"(?P<store>\d+)\s+"
    r"(?P<qty>\d+)\s+"
    r"(?P<sku>\d{5,})\s+"
    r"(?P<desc>.+?)\s+"
    r"(?P<size>\d{2}X\d{2})\s+"
    r"(?P<mfg>[A-Z0-9\-]+)\s+"
    r"(?P<unit_cost>\d+\.\d{2})\s+"
    r"(?P<uom>[A-Z]{2})\s+"
    r"(?P<ext_cost>\d+\.\d{2})\s*$"
)

PO_NUMBER_PATTERN = re.compile(r"P\s*\.?\s*O\s*\.?\s*#\s*:?\s*([\d\s]+)")
CLIENT_NAME_PATTERN = re.compile(r"^(.+?)\s*\nPage:", re.MULTILINE)


def extract_po_metadata(full_text: str) -> dict:
    """Extrae metadatos generales de la orden (PO#, cliente, fechas, total)."""
    meta = {}

    po_match = PO_NUMBER_PATTERN.search(full_text)
    if po_match:
        meta["po_number"] = po_match.group(1).replace(" ", "").strip()

    lines = full_text.splitlines()
    for line in lines:
        stripped = line.strip()
        if "Page:" in stripped:
            meta["client_name"] = stripped.split("Page:")[0].strip()
            break

    date_match = re.search(r"Order Date:\s*([\d/\s]+)", full_text)
    if date_match:
        meta["order_date"] = date_match.group(1).strip()

    total_units_match = re.search(r"TOTAL UNITS\s+(\d+)", full_text)
    if total_units_match:
        meta["total_units_pdf"] = int(total_units_match.group(1))

    total_cost_match = re.search(r"TOTAL P\.O\.\s+([\d,]+\.\d{2})", full_text)
    if total_cost_match:
        meta["total_cost_pdf"] = float(total_cost_match.group(1).replace(",", ""))

    return meta


def extract_lines_from_pdf(file) -> tuple[list[dict], dict, list[str]]:
    """
    Lee el PDF y devuelve:
      - lista de líneas de detalle parseadas
      - metadatos generales de la orden
      - lista de advertencias (líneas que no se pudieron interpretar)
    """
    rows = []
    warnings = []
    full_text_pages = []

    with pdfplumber.open(file) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            full_text_pages.append(text)

            for raw_line in text.splitlines():
                line = raw_line.strip()
                if not line or not line[0].isdigit():
                    continue

                match = LINE_PATTERN.match(line)
                if match:
                    d = match.groupdict()
                    cintura, largo = d["size"].split("X")
                    rows.append({
                        "page": page_num,
                        "line_no": int(d["line_no"]),
                        "store": int(d["store"]),
                        "qty": int(d["qty"]),
                        "sku": d["sku"],
                        "desc": d["desc"].strip(),
                        "size": d["size"],
                        "cintura": int(cintura),
                        "largo": int(largo),
                        "mfg": d["mfg"],
                        "unit_cost": float(d["unit_cost"]),
                        "ext_cost": float(d["ext_cost"]),
                    })
                else:
                    # Solo marcar como advertencia si parece una línea de detalle
                    # (empieza con número y tiene varios tokens) pero no calzó el patrón
                    tokens = line.split()
                    if len(tokens) >= 6 and tokens[0].isdigit():
                        warnings.append(f"Página {page_num}: no se pudo interpretar -> {line}")

    full_text = "\n".join(full_text_pages)
    meta = extract_po_metadata(full_text)
    return rows, meta, warnings


# ----------------------------------------------------------------------------
# AGRUPACIÓN Y VALIDACIÓN
# ----------------------------------------------------------------------------

def detect_description_mfg_mismatches(rows: list[dict]) -> list[dict]:
    """
    Detecta líneas donde la descripción (estilo/lavado en texto) no concuerda
    con el grupo dominante de descripciones para ese código MFG.
    Devuelve una lista de inconsistencias para mostrarle al usuario.
    """
    mfg_to_descs = defaultdict(lambda: defaultdict(int))
    for r in rows:
        mfg_to_descs[r["mfg"]][r["desc"]] += 1

    # Descripción dominante por MFG
    dominant_desc = {
        mfg: max(descs.items(), key=lambda kv: kv[1])[0]
        for mfg, descs in mfg_to_descs.items()
    }

    mismatches = []
    for r in rows:
        if r["desc"] != dominant_desc[r["mfg"]]:
            mismatches.append({
                "line_no": r["line_no"],
                "sku": r["sku"],
                "size": r["size"],
                "desc_en_linea": r["desc"],
                "mfg": r["mfg"],
                "desc_esperada_por_mfg": dominant_desc[r["mfg"]],
            })
    return mismatches


def group_by_mfg(rows: list[dict]) -> dict:
    """Agrupa las líneas por código MFG (estilo + lavado de fabricación)."""
    groups = defaultdict(lambda: {"rows": [], "descs": defaultdict(int)})
    for r in rows:
        groups[r["mfg"]]["rows"].append(r)
        groups[r["mfg"]]["descs"][r["desc"]] += 1
    return groups


def build_size_matrix(rows: list[dict]) -> tuple[list[int], list[int], dict]:
    """Construye la matriz cintura x largo -> cantidad para un grupo de líneas."""
    cinturas = sorted(set(r["cintura"] for r in rows))
    largos = sorted(set(r["largo"] for r in rows))
    qty_map = defaultdict(int)
    for r in rows:
        qty_map[(r["cintura"], r["largo"])] += r["qty"]
    return cinturas, largos, qty_map


# Sufijos de lavado conocidos en las descripciones de Big R (ajustar si aparecen nuevos)
_KNOWN_WASH_SUFFIXES = ["MSH", "RSH", "MSTN", "LSTN"]


def split_style_wash(desc: str) -> tuple[str, str]:
    """
    Separa una descripción de texto (p. ej. 'M 5 PCKT JEAN MSH') en
    (estilo, lavado), usando los sufijos de lavado conocidos.
    Si no se reconoce ningún sufijo, devuelve la descripción completa como
    estilo y 'N/D' como lavado.
    """
    tokens = desc.strip().split()
    if tokens and tokens[-1] in _KNOWN_WASH_SUFFIXES:
        return " ".join(tokens[:-1]), tokens[-1]
    return desc.strip(), "N/D"


# ----------------------------------------------------------------------------
# GENERACIÓN DEL EXCEL
# ----------------------------------------------------------------------------

# Estilos reutilizables
_TITLE_FONT = Font(name="Arial", size=14, bold=True)
_SUBTITLE_FONT = Font(name="Arial", size=10, italic=True)
_HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", start_color="2F5496")
_MFG_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
_MFG_FILL = PatternFill("solid", start_color="595959")
_TOTAL_FONT = Font(name="Arial", size=10, bold=True)
_TOTAL_FILL = PatternFill("solid", start_color="D9D9D9")
_NORMAL_FONT = Font(name="Arial", size=10)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_PO_HEADER_FILL = PatternFill("solid", start_color="4472C4")


def _write_matrix_sheet(ms, title: str, subtitle: str, rows: list[dict]) -> None:
    """Escribe una matriz Cintura x Largo en una hoja ya creada."""
    ms["A1"] = title
    ms["A1"].font = _TITLE_FONT
    ms["A2"] = subtitle
    ms["A2"].font = _SUBTITLE_FONT

    cinturas, largos, qty_map = build_size_matrix(rows)

    start_row = 4
    ms.cell(row=start_row, column=1, value="LARGO \\ CINTURA").font = _HEADER_FONT
    ms.cell(row=start_row, column=1).fill = _HEADER_FILL
    ms.cell(row=start_row, column=1).border = _BORDER
    ms.cell(row=start_row, column=1).alignment = _CENTER

    for j, cintura in enumerate(cinturas, start=2):
        cell = ms.cell(row=start_row, column=j, value=cintura)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = _CENTER

    total_col = len(cinturas) + 2
    cell = ms.cell(row=start_row, column=total_col, value="TOTAL")
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.border = _BORDER
    cell.alignment = _CENTER

    for i, largo in enumerate(largos, start=1):
        r = start_row + i
        cell = ms.cell(row=r, column=1, value=largo)
        cell.font = _MFG_FONT
        cell.fill = _MFG_FILL
        cell.border = _BORDER
        cell.alignment = _CENTER

        for j, cintura in enumerate(cinturas, start=2):
            qty = qty_map.get((cintura, largo), 0)
            c = ms.cell(row=r, column=j)
            c.border = _BORDER
            c.alignment = _CENTER
            c.font = _NORMAL_FONT
            c.value = qty if qty else None

        col_start = get_column_letter(2)
        col_end = get_column_letter(total_col - 1)
        c = ms.cell(row=r, column=total_col, value=f"=SUM({col_start}{r}:{col_end}{r})")
        c.font = _TOTAL_FONT
        c.fill = _TOTAL_FILL
        c.border = _BORDER
        c.alignment = _CENTER

    total_row = start_row + len(largos) + 1
    cell = ms.cell(row=total_row, column=1, value="TOTAL")
    cell.font = _TOTAL_FONT
    cell.fill = _TOTAL_FILL
    cell.border = _BORDER
    cell.alignment = _CENTER

    for j, cintura in enumerate(cinturas, start=2):
        col_letter = get_column_letter(j)
        c = ms.cell(
            row=total_row, column=j,
            value=f"=SUM({col_letter}{start_row + 1}:{col_letter}{start_row + len(largos)})"
        )
        c.font = _TOTAL_FONT
        c.fill = _TOTAL_FILL
        c.border = _BORDER
        c.alignment = _CENTER

    col_start = get_column_letter(2)
    col_end = get_column_letter(total_col - 1)
    c = ms.cell(row=total_row, column=total_col, value=f"=SUM({col_start}{total_row}:{col_end}{total_row})")
    c.font = _TOTAL_FONT
    c.fill = _TOTAL_FILL
    c.border = _BORDER
    c.alignment = _CENTER

    ms.column_dimensions["A"].width = 18
    for j in range(2, total_col + 1):
        ms.column_dimensions[get_column_letter(j)].width = 9


def _unique_sheet_name(wb, base: str) -> str:
    """Genera un nombre de hoja único y válido para Excel (máx 31 caracteres)."""
    name = re.sub(r"[^A-Za-z0-9\-_]", "", base)[:31] or "Hoja"
    candidate = name
    n = 1
    while candidate in wb.sheetnames:
        suffix = f"_{n}"
        candidate = name[: 31 - len(suffix)] + suffix
        n += 1
    return candidate


def build_excel_multi(orders: list[dict]) -> bytes:
    """
    Genera el Excel consolidado a partir de una lista de órdenes ya procesadas.
    Cada elemento de `orders` es un dict con: meta, rows, groups, total_qty, mismatches.

    Estructura del archivo:
      - Resumen global   -> totales por PO y por estilo/lavado, todo en una tabla plana
      - Validación       -> control de calidad por PDF (líneas leídas, piezas, advertencias)
      - Global_<MFG>     -> una matriz Cintura x Largo por estilo/lavado, sumando TODAS las PO
      - <PO>_<MFG>       -> una matriz Cintura x Largo por estilo/lavado, de cada PO individual
    """
    wb = Workbook()

    # ---------------- Hoja: Resumen global (tabla plana PO|Cliente|Estilo|Lavado|Cantidad) ----------------
    ws = wb.active
    ws.title = "Resumen global"

    ws["A1"] = "RESUMEN GLOBAL - TODAS LAS ÓRDENES DE COMPRA"
    ws["A1"].font = _TITLE_FONT
    total_general = sum(o["total_qty"] for o in orders)
    ws["A2"] = f"PDFs procesados: {len(orders)}    |    Total piezas (todas las PO): {total_general}"
    ws["A2"].font = _SUBTITLE_FONT

    headers = ["PO", "Cliente", "Fecha orden", "Estilo", "Lavado", "Código MFG", "Cantidad"]
    row = 4
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.border = _BORDER
        c.alignment = _CENTER
    row += 1

    for o in orders:
        meta = o["meta"]
        po = meta.get("po_number", "N/D")
        client = meta.get("client_name", "N/D")
        order_date = meta.get("order_date", "N/D")
        for mfg, g in o["groups"].items():
            qty = sum(r["qty"] for r in g["rows"])
            desc = max(g["descs"].items(), key=lambda kv: kv[1])[0]
            estilo, lavado = split_style_wash(desc)
            values = [po, client, order_date, estilo, lavado, mfg, qty]
            for j, v in enumerate(values, start=1):
                c = ws.cell(row=row, column=j, value=v)
                c.font = _NORMAL_FONT
                c.border = _BORDER
                if j == 7:
                    c.alignment = _CENTER
            row += 1

    # Fila de total general
    c = ws.cell(row=row, column=1, value="TOTAL GENERAL")
    c.font = _TOTAL_FONT
    c.fill = _TOTAL_FILL
    for j in range(2, 7):
        ws.cell(row=row, column=j).fill = _TOTAL_FILL
    c = ws.cell(row=row, column=7, value=total_general)
    c.font = _TOTAL_FONT
    c.fill = _TOTAL_FILL
    c.alignment = _CENTER

    col_widths = [14, 22, 13, 20, 22, 18, 11]
    for j, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # ---------------- Hoja: Validación (control de calidad por PDF) ----------------
    wv = wb.create_sheet("Validación")
    wv["A1"] = "VALIDACIÓN DE LECTURA POR ARCHIVO"
    wv["A1"].font = _TITLE_FONT

    headers_v = ["PO", "Cliente", "Líneas leídas", "Piezas extraídas", "Piezas según PDF", "¿Coincide?", "Advertencias"]
    row = 3
    for j, h in enumerate(headers_v, start=1):
        c = wv.cell(row=row, column=j, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.border = _BORDER
        c.alignment = _CENTER
    row += 1

    for o in orders:
        meta = o["meta"]
        po = meta.get("po_number", "N/D")
        client = meta.get("client_name", "N/D")
        n_lines = len(o["rows"])
        extracted = o["total_qty"]
        declared = meta.get("total_units_pdf")
        coincide = "SÍ" if declared is not None and declared == extracted else ("N/D" if declared is None else "NO")
        n_warn = len(o["mismatches"]) + len(o.get("parse_warnings", []))

        values = [po, client, n_lines, extracted, declared if declared is not None else "N/D", coincide, n_warn]
        for j, v in enumerate(values, start=1):
            c = wv.cell(row=row, column=j, value=v)
            c.font = _NORMAL_FONT
            c.border = _BORDER
            if j in (3, 4, 5, 7):
                c.alignment = _CENTER
            if j == 6:
                c.alignment = _CENTER
                if coincide == "NO":
                    c.font = Font(name="Arial", size=10, bold=True, color="C00000")
        row += 1

    # Detalle de inconsistencias descripción/MFG
    row += 1
    wv.cell(row=row, column=1, value="DETALLE DE INCONSISTENCIAS DESCRIPCIÓN / CÓDIGO MFG").font = Font(
        name="Arial", size=11, bold=True
    )
    row += 1
    headers_m = ["PO", "Línea", "SKU", "Talla", "Descripción en línea", "Código MFG", "Descripción esperada"]
    for j, h in enumerate(headers_m, start=1):
        c = wv.cell(row=row, column=j, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.border = _BORDER
    row += 1
    for o in orders:
        po = o["meta"].get("po_number", "N/D")
        for m in o["mismatches"]:
            values = [po, m["line_no"], m["sku"], m["size"], m["desc_en_linea"], m["mfg"], m["desc_esperada_por_mfg"]]
            for j, v in enumerate(values, start=1):
                c = wv.cell(row=row, column=j, value=v)
                c.font = _NORMAL_FONT
                c.border = _BORDER
            row += 1

    col_widths_v = [14, 22, 13, 15, 16, 12, 13]
    for j, w in enumerate(col_widths_v, start=1):
        wv.column_dimensions[get_column_letter(j)].width = w

    # ---------------- Hojas: Matriz global por estilo/lavado (todas las PO sumadas) ----------------
    global_rows_by_mfg = defaultdict(list)
    global_descs_by_mfg = defaultdict(lambda: defaultdict(int))
    for o in orders:
        for mfg, g in o["groups"].items():
            global_rows_by_mfg[mfg].extend(g["rows"])
            for desc, count in g["descs"].items():
                global_descs_by_mfg[mfg][desc] += count

    for mfg, rows in global_rows_by_mfg.items():
        desc = max(global_descs_by_mfg[mfg].items(), key=lambda kv: kv[1])[0]
        total = sum(r["qty"] for r in rows)
        sheet_name = _unique_sheet_name(wb, f"GLOBAL_{mfg}")
        ms = wb.create_sheet(sheet_name)
        _write_matrix_sheet(
            ms,
            title=f"{desc} — TODAS LAS ÓRDENES",
            subtitle=f"Código MFG: {mfg}    |    Total piezas (global): {total}    |    PDFs incluidos: {len(orders)}",
            rows=rows,
        )

    # ---------------- Hojas: Matriz por PO individual ----------------
    for o in orders:
        meta = o["meta"]
        po = meta.get("po_number", "N/D")
        for mfg, g in o["groups"].items():
            desc = max(g["descs"].items(), key=lambda kv: kv[1])[0]
            total = sum(r["qty"] for r in g["rows"])
            sheet_name = _unique_sheet_name(wb, f"{po}_{mfg}")
            ms = wb.create_sheet(sheet_name)
            _write_matrix_sheet(
                ms,
                title=f"{desc} — PO {po}",
                subtitle=f"Cliente: {meta.get('client_name', 'N/D')}    |    Código MFG: {mfg}    |    Total piezas: {total}",
                rows=g["rows"],
            )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ----------------------------------------------------------------------------
# INTERFAZ STREAMLIT
# ----------------------------------------------------------------------------

import pandas as pd

st.set_page_config(page_title="Matriz de Producción desde PO", layout="wide")
st.title("📋 Generador de Matriz de Producción")
st.caption(
    "Sube una o varias órdenes de compra en PDF (formato Big R) y obtén el resumen, "
    "la validación de lectura y las matrices Cintura x Largo por estilo/lavado."
)

uploaded_files = st.file_uploader(
    "Selecciona uno o varios PDFs de órdenes de compra",
    type=["pdf"],
    accept_multiple_files=True,
)

if uploaded_files:
    orders = []
    failed_files = []

    with st.spinner(f"Leyendo y procesando {len(uploaded_files)} archivo(s)..."):
        for uf in uploaded_files:
            rows, meta, parse_warnings = extract_lines_from_pdf(uf)
            if not rows:
                failed_files.append(uf.name)
                continue
            total_qty = sum(r["qty"] for r in rows)
            mismatches = detect_description_mfg_mismatches(rows)
            groups = group_by_mfg(rows)
            orders.append({
                "file_name": uf.name,
                "meta": meta,
                "rows": rows,
                "total_qty": total_qty,
                "mismatches": mismatches,
                "groups": groups,
                "parse_warnings": parse_warnings,
            })

    if failed_files:
        st.error(
            "No se pudo extraer ninguna línea de detalle de: " + ", ".join(failed_files) +
            ". Es posible que el formato sea distinto al esperado."
        )

    if not orders:
        st.stop()

    total_general = sum(o["total_qty"] for o in orders)
    total_lineas = sum(len(o["rows"]) for o in orders)
    total_advertencias = sum(len(o["mismatches"]) + len(o["parse_warnings"]) for o in orders)

    # ---------------- Panel de validación consolidado ----------------
    st.subheader("Validación de lectura")
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("PDFs procesados", len(orders))
    col2.metric("Líneas leídas", total_lineas)

    declared_total = sum(
        o["meta"]["total_units_pdf"] for o in orders if o["meta"].get("total_units_pdf") is not None
    )
    if declared_total:
        col3.metric(
            "Piezas validadas",
            f"{total_general}/{declared_total}",
            delta="OK" if total_general == declared_total else "Revisar",
            delta_color="normal" if total_general == declared_total else "inverse",
        )
    else:
        col3.metric("Piezas extraídas", total_general)
    col4.metric("Advertencias", total_advertencias)

    # Tabla de validación por archivo
    val_rows = []
    for o in orders:
        meta = o["meta"]
        declared = meta.get("total_units_pdf")
        coincide = "✅" if declared is not None and declared == o["total_qty"] else ("N/D" if declared is None else "⚠️")
        val_rows.append({
            "Archivo": o["file_name"],
            "PO": meta.get("po_number", "N/D"),
            "Cliente": meta.get("client_name", "N/D"),
            "Líneas leídas": len(o["rows"]),
            "Piezas extraídas": o["total_qty"],
            "Piezas según PDF": declared if declared is not None else "N/D",
            "¿Coincide?": coincide,
            "Advertencias": len(o["mismatches"]) + len(o["parse_warnings"]),
        })
    st.dataframe(pd.DataFrame(val_rows), use_container_width=True, hide_index=True)

    # Advertencias de líneas no interpretadas
    all_parse_warnings = [(o["file_name"], w) for o in orders for w in o["parse_warnings"]]
    if all_parse_warnings:
        with st.expander(f"⚠️ {len(all_parse_warnings)} línea(s) no se pudieron interpretar automáticamente"):
            for fname, w in all_parse_warnings:
                st.text(f"[{fname}] {w}")

    # Inconsistencias descripción vs código MFG
    all_mismatches = [(o["meta"].get("po_number", "N/D"), m) for o in orders for m in o["mismatches"]]
    if all_mismatches:
        with st.expander(f"⚠️ {len(all_mismatches)} línea(s) con posible inconsistencia descripción/código MFG"):
            st.write(
                "Estas líneas tienen una descripción de estilo/lavado distinta a la dominante "
                "para su código MFG. Se agruparon por código MFG, pero conviene confirmar con el cliente."
            )
            for po, m in all_mismatches:
                st.text(
                    f"PO {po} | Línea {m['line_no']} | SKU {m['sku']} | Talla {m['size']} | "
                    f"MFG: {m['mfg']} | Descripción en línea: '{m['desc_en_linea']}' | "
                    f"Descripción esperada: '{m['desc_esperada_por_mfg']}'"
                )

    # ---------------- Pestañas de resultados ----------------
    tab_resumen, tab_por_po, tab_global = st.tabs(["📋 Resumen simple", "📦 Matrices por PO", "🌐 Matriz global"])

    with tab_resumen:
        st.caption("Cantidad por PO, estilo y lavado.")
        resumen_rows = []
        for o in orders:
            meta = o["meta"]
            po = meta.get("po_number", "N/D")
            client = meta.get("client_name", "N/D")
            for mfg, g in o["groups"].items():
                qty = sum(r["qty"] for r in g["rows"])
                desc = max(g["descs"].items(), key=lambda kv: kv[1])[0]
                estilo, lavado = split_style_wash(desc)
                resumen_rows.append({
                    "PO": po, "Cliente": client, "Estilo": estilo, "Lavado": lavado, "Cantidad": qty,
                })
        df_resumen = pd.DataFrame(resumen_rows)
        st.dataframe(df_resumen, use_container_width=True, hide_index=True)
        st.markdown(f"**Total general: {total_general} piezas**")

    with tab_por_po:
        st.caption("Una matriz Cintura x Largo por cada estilo/lavado, separada por orden de compra.")
        po_options = [o["meta"].get("po_number", o["file_name"]) for o in orders]
        selected_po = st.selectbox("Selecciona la orden de compra", po_options)
        selected_order = next(
            o for o in orders if o["meta"].get("po_number", o["file_name"]) == selected_po
        )
        st.markdown(
            f"**Cliente:** {selected_order['meta'].get('client_name', 'N/D')}  |  "
            f"**Fecha orden:** {selected_order['meta'].get('order_date', 'N/D')}  |  "
            f"**Total piezas:** {selected_order['total_qty']}"
        )
        for mfg, g in selected_order["groups"].items():
            desc = max(g["descs"].items(), key=lambda kv: kv[1])[0]
            total = sum(r["qty"] for r in g["rows"])
            cinturas, largos, qty_map = build_size_matrix(g["rows"])
            st.markdown(f"**{desc}** — `{mfg}` — Total: {total} pzs")
            table = pd.DataFrame(
                [[qty_map.get((c, l), 0) or "" for c in cinturas] for l in largos],
                index=[f"Largo {l}" for l in largos],
                columns=[f"Cintura {c}" for c in cinturas],
            )
            st.dataframe(table, use_container_width=True)

    with tab_global:
        st.caption("Una matriz Cintura x Largo por cada estilo/lavado, sumando todas las órdenes de compra cargadas.")
        global_rows_by_mfg = defaultdict(list)
        global_descs_by_mfg = defaultdict(lambda: defaultdict(int))
        for o in orders:
            for mfg, g in o["groups"].items():
                global_rows_by_mfg[mfg].extend(g["rows"])
                for desc, count in g["descs"].items():
                    global_descs_by_mfg[mfg][desc] += count

        for mfg, rows in global_rows_by_mfg.items():
            desc = max(global_descs_by_mfg[mfg].items(), key=lambda kv: kv[1])[0]
            total = sum(r["qty"] for r in rows)
            cinturas, largos, qty_map = build_size_matrix(rows)
            st.markdown(f"**{desc}** — `{mfg}` — Total: {total} pzs (todas las PO)")
            table = pd.DataFrame(
                [[qty_map.get((c, l), 0) or "" for c in cinturas] for l in largos],
                index=[f"Largo {l}" for l in largos],
                columns=[f"Cintura {c}" for c in cinturas],
            )
            st.dataframe(table, use_container_width=True)

    # ---------------- Descarga ----------------
    st.subheader("Descargar resultado")
    excel_bytes = build_excel_multi(orders)
    if len(orders) == 1:
        file_name = f"Matriz_PO_{orders[0]['meta'].get('po_number', 'orden')}.xlsx"
    else:
        file_name = f"Matriz_consolidada_{len(orders)}_PO.xlsx"
    st.download_button(
        label="⬇️ Descargar Excel consolidado",
        data=excel_bytes,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.caption(
        "El Excel incluye: Resumen global, Validación, una hoja Global por estilo/lavado "
        "(todas las PO sumadas) y una hoja por estilo/lavado de cada PO individual."
    )
else:
    st.info("Sube uno o varios PDFs para comenzar.")
